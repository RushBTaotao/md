"""
处理包含PMF数据的Excel文件并生成甘特图和汇总图。

此脚本读取Excel文件('mrg.xlsx')，提取以'PMF'开头的sheet，
保存为CSV，并生成单个甘特PNG，
收集PMF任务，清理和处理它们，并生成按size和round分组的汇总图。
"""

import openpyxl
import csv
import sys
import os
import matplotlib
matplotlib.use('Agg') # 使用非交互式后端
import matplotlib.pyplot as plt
from collections import defaultdict
import re
from concurrent.futures import ProcessPoolExecutor
import gantt_scheduler  # 导入重构后的绘图模块
import time

def read_tasks_from_csv(csv_file_path):
    """
    从CSV文件中读取任务和配置。
    """
    return gantt_scheduler.read_tasks(csv_file_path)

def get_round(sheet):
    """
    从sheet名称确定round。
    """
    if 'round0' in sheet and 'round0-3' not in sheet:
        return '0'
    else:
        return '1'

def get_c(sheet):
    """
    从sheet名称确定c。
    """
    if 'c0' in sheet:
        return 'c0'
    elif 'c1' in sheet:
        return 'c1'
    else:
        return 'other'

def clean_pmf_tasks(tasks):
    """
    清理和调整PMF任务。
    """
    cleaned = []
    for task in tasks:
        mode = task['mode']
        size = get_size(mode)
        uv = task['uv']

        # 1. Remove _a/_b for specific conditions
        remove = False
        if uv == 'Y' and size in ['16', '32', '64'] and ('_a' in task.get('original_mode', mode) or '_b' in task.get('original_mode', mode)):
            remove = True
        if uv == 'UV' and size in ['8', '16', '32'] and ('_a' in task.get('original_mode', mode) or '_b' in task.get('original_mode', mode)):
            remove = True
        if remove:
            continue

        cleaned.append({
            'mode': mode,
            'output_begin': task['output_begin'],
            'output_end': task['output_end'],
            'sheet': task['sheet'],
            'round': task['round'],
            'c': task['c'],
            'uv': task['uv']
        })
    return cleaned

def get_size(mode):
    match = re.search(r'[MF](\d+)', mode)
    if match:
        return match.group(1)
    else:
        return 'other'

def find_overlaps(intervals):
    if not intervals:
        return []
    overlaps = []
    for i in range(len(intervals)):
        for j in range(i+1, len(intervals)):
            start1, end1 = intervals[i]
            start2, end2 = intervals[j]
            overlap_start = max(start1, start2)
            overlap_end = min(end1, end2)
            if overlap_start < overlap_end:
                overlaps.append((overlap_start, overlap_end))
    return overlaps

def plot_single_summary(grouped, filename, title, xlim=None):
    plt.figure(figsize=(19, 10))
    plt.clf()

    y_pos = 0
    labels = []
    all_times = []
    for (size, uv), task_list in grouped.items():
        if xlim:
            task_list = [task for task in task_list if task['output_begin'] is not None and task['output_end'] is not None and task['output_begin'] >= xlim[0] and task['output_end'] <= xlim[1]]
        if not task_list:
            continue
        labels.append(f'{size} {uv}')
        for task in task_list:
            ob = task['output_begin']
            oe = task['output_end']
            if ob is not None and oe is not None:
                duration = oe - ob
                if duration > 0:
                    color = 'moccasin' if uv == 'Y' else 'lightgreen'
                    plt.barh(y_pos, duration, left=ob, height=0.4, color=color)
                    mode_parts = task['mode'].split('_')
                    if 'sp' in mode_parts:
                        suffix_parts = []
                        for p in mode_parts[1:]:
                            if p in ['Y', 'UV']:
                                break
                            suffix_parts.append(p)
                        mode_short = '_'.join(suffix_parts)
                    elif len(mode_parts) > 6:
                        mode_short = '_'.join(mode_parts[1:4])
                    else:
                        mode_short = '_'.join(mode_parts[1:3])
                    
                    suffix = f"{task['uv']} {task['c']} {mode_short}"
                    plt.text(ob + duration / 2, y_pos, suffix, ha='center', va='center', fontsize=7, color='black', weight='bold', rotation=45)
                all_times.extend([ob, oe])
        overlaps = find_overlaps([(task['output_begin'], task['output_end']) for task in task_list if task['output_begin'] is not None and task['output_end'] is not None])
        if overlaps:
            broken_overlaps = [(s, e - s) for s, e in overlaps]
            plt.broken_barh(broken_overlaps, (y_pos, 0.4), facecolors='red')
        y_pos += 1

    all_tasks_in_group = [task for task_list in grouped.values() for task in task_list]
    if all_tasks_in_group:
        for task in all_tasks_in_group:
            ob = task['output_begin']
            oe = task['output_end']
            if ob is not None and oe is not None:
                duration = oe - ob
                if duration > 0:
                    plt.barh(y_pos, duration, left=ob, height=0.4, color='gray')
        overlaps = find_overlaps([(task['output_begin'], task['output_end']) for task in all_tasks_in_group if task['output_begin'] is not None and task['output_end'] is not None])
        if overlaps:
            broken_overlaps = [(s, e - s) for s, e in overlaps]
            plt.broken_barh(broken_overlaps, (y_pos, 0.4), facecolors='red')
        all_times_sum = [t for task in all_tasks_in_group for t in [task['output_begin'], task['output_end']] if t is not None]
        if all_times_sum:
            total_start = min(all_times_sum)
            total_end = max(all_times_sum)
            plt.text((total_start + total_end) / 2, y_pos, 'Summary', ha='center', va='center', fontsize=7, color='black', weight='bold')
        labels.append('Summary')
        y_pos += 1

    plt.yticks(range(len(labels)), labels)
    plt.xlabel('Clock Cycles')
    plt.ylabel('Size / Type / C')
    plt.title(title)

    tick_positions = set([task['output_begin'] for task in all_tasks_in_group if task['output_begin'] is not None])
    if '16' in filename or '32' in filename:
        tick_positions.update([task['output_end'] for task in all_tasks_in_group if task['output_end'] is not None])
    tick_positions = sorted(list(tick_positions))
    if tick_positions:
        plt.xticks(tick_positions, [str(t) for t in tick_positions], rotation=45, ha='right')

    if all_times:
        min_t = min(all_times)
        max_t = max(all_times)
        if '8_round1' in filename:
            plt.xlim(max(0, min_t - 2), 400)
        else:
            plt.xlim(max(0, min_t - 2), max_t)
    elif xlim:
        plt.xlim(xlim)

    plt.grid(True, axis='x')
    if os.path.exists(filename):
        os.remove(filename)
    if '16' in filename or '32' in filename:
        plt.savefig(filename, dpi=300)
    else:
        plt.savefig(filename, dpi=300, bbox_inches='tight')
    plt.close()

def generate_summary_plot(tasks, sizes, r, xlim):
    filtered = [task for task in tasks if get_size(task['mode']) in sizes and task['round'] == r]
    if '8' in sizes and r != '1':
        filtered = [task for task in filtered if task['output_end'] is None or task['output_end'] <= 200]
    
    if not filtered: return 0

    valid_ob = [task['output_begin'] for task in filtered if task['output_begin'] is not None]
    if not valid_ob: return 0
    min_ob = min(valid_ob)
    
    # Special handling for round 1 xlim
    if r == '1':
        if '4' in sizes:
            valid_oe = [task['output_end'] for task in filtered if task['output_end'] is not None]
            max_oe = max(valid_oe) if valid_oe else 0
            xlim = (min_ob, max_oe)
        elif '8' in sizes:
            xlim = (min_ob if min_ob < 0 else 0, 400)
    else:
        if min_ob < xlim[0]:
            xlim = (min_ob, xlim[1])

    grouped = defaultdict(list)
    for task in filtered:
        size = get_size(task['mode'])
        uv = task['uv']
        grouped[(size, uv)].append(task)
    
    if grouped:
        size_str = '_'.join(sizes)
        plot_single_summary(grouped, f'PMF_Summary_{size_str}_round{r}.png', f'PMF Output Summary {"/".join(sizes)} Round {r}', xlim)
        return 1
    return 0

def generate_combined_summary_plot(tasks, size, xlim):
    filtered = [task for task in tasks if get_size(task['mode']) == size]
    if filtered:
        grouped = defaultdict(list)
        for task in filtered:
            uv = task['uv']
            grouped[(size, uv)].append(task)
        if grouped:
            plot_single_summary(grouped, f'PMF_Summary_{size}.png', f'PMF Output Summary {size}', xlim)
            return 1
    return 0

def process_single_sheet_data(sheet_data, sheet_name, output_dir):
    """
    处理单个 Sheet 数据：存 CSV，绘图。
    作为并行执行的基础单元。
    """
    csv_file = os.path.join(output_dir, f"{sheet_name}.csv")
    png_file = os.path.join(output_dir, f"{sheet_name}.png")
    
    # Save to CSV
    try:
        with open(csv_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(sheet_data)
    except Exception as e:
        return f"Error saving {sheet_name}: {str(e)}", None, None

    png_generated = False
    # If not sp sheet, generate PNG
    if 'sp' not in sheet_name:
        try:
            tasks, config = gantt_scheduler.read_tasks(csv_file)
            if tasks:
                gantt_scheduler.plot_gantt(tasks, config, output=png_file, save_only=True)
                png_generated = True
        except Exception as e:
            return f"Error plotting {sheet_name}: {str(e)}", csv_file, False
    
    return None, csv_file, png_generated

def process_category(wb, sheets, category_name, output_dir):
    if not sheets:
        return [], {}

    os.makedirs(output_dir, exist_ok=True)
    print(f"Processing category {category_name} in {output_dir}")

    # Results tracking
    stats = {
        'sheets_found': len(sheets),
        'csv_created': 0,
        'png_created': 0,
        'summary_created': 0,
        'errors': []
    }

    # Step 1: Prepare data for all sheets
    sheet_data_list = []
    for sheet_name in sheets:
        sheet = wb[sheet_name]
        data = [[cell.value for cell in row] for row in sheet.rows]
        sheet_data_list.append((data, sheet_name, output_dir))

    # Step 2: Parallel process sheets (CSV and PNG)
    with ProcessPoolExecutor() as executor:
        futures = [executor.submit(process_single_sheet_data, *args) for args in sheet_data_list]
        for future in futures:
            err, csv_path, png_gen = future.result()
            if err:
                stats['errors'].append(err)
            if csv_path:
                stats['csv_created'] += 1
            if png_gen:
                stats['png_created'] += 1

    # Step 3: Collect tasks for summary
    category_tasks = []
    for sheet_name in sheets:
        csv_file = os.path.join(output_dir, f"{sheet_name}.csv")
        if not os.path.exists(csv_file): continue

        try:
            if 'sp' in sheet_name:
                tasks, _ = gantt_scheduler.read_tasks(csv_file)
                sp_tasks = [t for t in tasks if t['mode'].startswith('PMF_') and ('M8' in t['mode'] or 'F8' in t['mode'])]
                uv = 'Y'
                c_str = get_c(sheet_name)
                round_str = get_round(sheet_name)
                for t in sp_tasks:
                    if 'M8' in t['mode'] or 'F8' in t['mode']:
                        t['mode'] = t['mode'].replace('M8', 'M4').replace('F8', 'F4')
                    t['pipe_begin'] = t['pipe_end'] = t['input_begin'] = t['input_end'] = None
                    parts = t['mode'].split('_')
                    if len(parts) > 1:
                        t['mode'] = f"PMF_sp_{'_'.join(parts[1:])}"
                    task = {**t, 'sheet': sheet_name, 'round': round_str, 'c': c_str, 'uv': uv, 'original_mode': t['mode']}
                    category_tasks.append(task)
            else:
                uv = 'UV' if 'UV' in sheet_name else 'Y'
                c_str = get_c(sheet_name)
                rounds = ['0', '1'] if 'round0-3' in sheet_name else [get_round(sheet_name)]
                tasks, _ = gantt_scheduler.read_tasks(csv_file)
                pmf_tasks = [t for t in tasks if t['mode'].startswith('PMF_')]
                for t in pmf_tasks:
                    for round_str in rounds:
                        task = {**t, 'sheet': sheet_name, 'round': round_str, 'c': c_str, 'uv': uv, 'original_mode': t['mode']}
                        task['mode'] = f"{t['mode']}_{uv}_{c_str}_{round_str}"
                        category_tasks.append(task)
        except Exception as e:
            stats['errors'].append(f"Error collecting tasks for {sheet_name}: {str(e)}")

    # Step 4: Clean and Plot Summaries
    cleaned_tasks = clean_pmf_tasks(category_tasks)
    orig_cwd = os.getcwd()
    os.chdir(output_dir)
    try:
        for size in ['16', '32']:
            stats['summary_created'] += generate_combined_summary_plot(cleaned_tasks, size, (0, 800))
        for size in ['4', '8']:
            for r in ['0', '1']:
                stats['summary_created'] += generate_summary_plot(cleaned_tasks, [size], r, (0, 200))
    except Exception as e:
        stats['errors'].append(f"Error generating summary plots: {str(e)}")
    finally:
        os.chdir(orig_cwd)

    return cleaned_tasks, stats

def main():
    start_time = time.time()
    excel_file = 'mrg.xlsx'
    if not os.path.exists(excel_file):
        print(f"Error: Excel file '{excel_file}' not found.")
        sys.exit(1)

    print(f"Loading workbook: {excel_file}...")
    wb = openpyxl.load_workbook(excel_file, data_only=True, read_only=True)
    
    pmf_sheets = [s for s in wb.sheetnames if s.startswith('PMF')]
    pmf264_sheets = [s for s in wb.sheetnames if s.startswith('264PMF')]

    all_stats = {}

    if not pmf_sheets and not pmf264_sheets:
        print("No matching sheets found.")
    else:
        if pmf_sheets:
            _, all_stats['PMF'] = process_category(wb, pmf_sheets, "PMF", "PMF_Output")
        if pmf264_sheets:
            _, all_stats['264PMF'] = process_category(wb, pmf264_sheets, "264PMF", "264PMF_Output")

    total_time = time.time() - start_time

    # Generate Final Report
    print("\n" + "="*50)
    print("                EXECUTION REPORT")
    print("="*50)
    print(f"Total Execution Time: {total_time:.2f} seconds")
    
    total_found = 0
    total_png = 0
    total_summary = 0
    all_errors = []

    for cat, stats in all_stats.items():
        print(f"\n[{cat}] Category:")
        print(f"  - Sheets Processed: {stats['sheets_found']}")
        print(f"  - PNG Gantts Generated: {stats['png_created']}")
        print(f"  - Summary Plots Generated: {stats['summary_created']}")
        total_found += stats['sheets_found']
        total_png += stats['png_created']
        total_summary += stats['summary_created']
        all_errors.extend(stats['errors'])

    print("\n" + "-"*50)
    if not all_errors:
        print("Status: SUCCESS (No errors encountered)")
    else:
        print(f"Status: COMPLETED WITH {len(all_errors)} ERRORS")
        for i, error in enumerate(all_errors, 1):
            print(f"  {i}. {error}")
    
    print("\nOutput Directories:")
    print("  - PMF_Output/")
    print("  - 264PMF_Output/")
    print("="*50 + "\n")

if __name__ == "__main__":
    main()
